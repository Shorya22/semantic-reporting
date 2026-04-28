"""
Export service — CSV, Excel, and PDF report generation.

CSV:   pandas to_csv() — pure Python, UTF-8.
Excel: openpyxl workbook — styled data table + optional chart image sheet.
PDF:   fpdf2 — pure Python, no system dependencies.

All functions accept (rows, columns, **opts) and return raw bytes of the
target file format.
"""

from __future__ import annotations

import base64
import datetime as _dt
import io
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _strip_tz(v: Any) -> Any:
    """Excel/openpyxl can't write tz-aware datetimes — drop the tzinfo.

    PostgreSQL TIMESTAMPTZ columns flow back as ``datetime`` with ``tzinfo``
    attached, which makes ``openpyxl.cell._writer`` raise
    ``"Excel does not support timezones in datetimes"`` at save time.
    """
    if isinstance(v, _dt.datetime) and v.tzinfo is not None:
        return v.replace(tzinfo=None)
    if isinstance(v, _dt.time) and v.tzinfo is not None:
        return v.replace(tzinfo=None)
    return v


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def export_csv(rows: list, columns: list[str]) -> bytes:
    """Return UTF-8 CSV bytes for the given rows and column names."""
    if not rows:
        return (",".join(columns) + "\n").encode()
    df = _to_df(rows, columns)
    return df.to_csv(index=False).encode()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def export_excel(
    rows: list,
    columns: list[str],
    sheet_name: str = "Data",
    chart_b64: Optional[str] = None,
    chart_title: str = "Visualization",
) -> bytes:
    """
    Create an .xlsx workbook and return its raw bytes.

    Sheet 1 — styled data table with alternating row colors.
    Sheet 2 — embedded chart image (only when *chart_b64* is provided).
    """
    df = _to_df(rows, columns)

    wb  = Workbook()
    ws  = wb.active
    ws.title = sheet_name  # type: ignore[assignment]

    # ── Styles ──────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="6366F1")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="CBD5E1")
    cell_bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill("solid", fgColor="F1F5F9")

    # ── Header ──────────────────────────────────────────────────────────────
    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col_name))
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = cell_bdr

    ws.row_dimensions[1].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, val in enumerate(row, 1):
            # openpyxl rejects tz-aware datetimes; coerce here defensively
            cell = ws.cell(row=ri, column=ci, value=_strip_tz(val))
            cell.fill      = fill
            cell.border    = cell_bdr
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Column widths ────────────────────────────────────────────────────────
    for ci, col_name in enumerate(df.columns, 1):
        col_data   = df.iloc[:, ci - 1].astype(str)
        max_len    = max(len(str(col_name)), col_data.str.len().max() if not col_data.empty else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(int(max_len) + 4, 50)

    ws.freeze_panes        = "A2"
    ws.auto_filter.ref     = ws.dimensions

    # ── Chart sheet ──────────────────────────────────────────────────────────
    if chart_b64:
        ws2 = wb.create_sheet("Chart")
        ws2.cell(row=1, column=1, value=chart_title).font = Font(bold=True, size=14, color="6366F1")
        ws2.row_dimensions[1].height = 22

        img_bytes  = base64.b64decode(chart_b64)
        img_stream = io.BytesIO(img_bytes)
        xl_img     = XLImage(img_stream)
        # Scale to fit a wide spreadsheet view
        xl_img.width  = min(xl_img.width,  960)
        xl_img.height = min(xl_img.height, 540)
        ws2.add_image(xl_img, "A3")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def export_pdf(
    rows: list,
    columns: list[str],
    title: str = "Data Report",
    chart_b64: Optional[str] = None,
) -> bytes:
    """
    Generate a landscape-A4 PDF report and return its raw bytes.

    Page 1 — title, timestamp, and formatted data table.
    Page 2 — chart image (only when *chart_b64* is provided).
    """
    import datetime
    from fpdf import FPDF

    df = _to_df(rows, columns)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ── Title ─────────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(99, 102, 241)
    pdf.cell(0, 12, title, new_x="LMARGIN", new_y="NEXT", align="C")

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 116, 139)
    ts = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    pdf.cell(0, 6, f"Generated: {ts}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    if not df.empty:
        page_w  = pdf.w - pdf.l_margin - pdf.r_margin
        n_cols  = len(df.columns)
        col_w   = page_w / n_cols
        row_h   = 7

        # Header
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(99, 102, 241)
        pdf.set_text_color(255, 255, 255)
        for col_name in df.columns:
            pdf.cell(col_w, row_h + 1, str(col_name)[:22], border=1, align="C", fill=True)
        pdf.ln()

        # Rows
        pdf.set_font("Helvetica", "", 8)
        for ri, row in enumerate(df.itertuples(index=False)):
            if ri % 2 == 0:
                pdf.set_fill_color(248, 250, 252)
            else:
                pdf.set_fill_color(241, 245, 249)
            pdf.set_text_color(15, 23, 42)
            for val in row:
                txt = "" if val is None else str(val)[:22]
                pdf.cell(col_w, row_h, txt, border=1, fill=True)
            pdf.ln()

    # ── Chart page ────────────────────────────────────────────────────────────
    if chart_b64:
        pdf.add_page()

        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(99, 102, 241)
        pdf.cell(0, 10, "Visualization", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(4)

        img_bytes  = base64.b64decode(chart_b64)
        img_stream = io.BytesIO(img_bytes)
        page_w     = pdf.w - pdf.l_margin - pdf.r_margin
        pdf.image(img_stream, x=pdf.l_margin, y=pdf.get_y(), w=page_w)

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Internal helper
# ---------------------------------------------------------------------------

def _to_df(rows: list, columns: list[str]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=columns)
    if isinstance(rows[0], dict):
        return pd.DataFrame(rows)
    return pd.DataFrame(rows, columns=columns)
